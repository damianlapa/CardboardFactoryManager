# views.py
import calendar
import datetime

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import WorkStation, ProductionUnit


def _to_local_str(dt):
    """Zwraca 'YYYY-MM-DD HH:MM' dla dt; bezpieczne dla naive/aware/None."""
    if not dt:
        return ""
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_default_timezone())
    return timezone.localtime(dt).strftime('%Y-%m-%d %H:%M')


class WorkStationLossesXLSX(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def get(self, request, year, month, station_id):
        try:
            year = int(year)
            month = int(month)
            if not (1 <= month <= 12):
                raise ValueError
        except ValueError:
            raise Http404("Nieprawidłowa data.")

        station = get_object_or_404(WorkStation, pk=station_id)

        # Zakres miesiąca
        date_start = datetime.date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        date_end = datetime.date(year, month, last_day)

        period_start = datetime.datetime.combine(date_start, datetime.time.min)
        period_end = datetime.datetime.combine(date_end + datetime.timedelta(days=1), datetime.time.min)

        tz = timezone.get_default_timezone()
        if timezone.is_naive(period_start):
            period_start = timezone.make_aware(period_start, tz)
        if timezone.is_naive(period_end):
            period_end = timezone.make_aware(period_end, tz)

        # Jednostki w okresie
        units = (
            ProductionUnit.objects
            .select_related('production_order', 'work_station')
            .filter(work_station=station)
            .filter(start__lt=period_end, end__gte=period_start)
            .order_by('start')
        )

        # Budowa XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Straty"

        headers = [
            "Maszyna", "Data startu", "Data końca", "Zlecenie", "Sekwencja",
            "Plan [min]", "Rzecz. [min]", "Qty IN", "Qty OUT", "LOSS", "LOSS [%]"
        ]
        ws.append(headers)

        sum_in = 0
        sum_out = 0
        sum_plan_min = 0
        sum_real_min = 0

        for u in units:
            planned_min = u.estimated_time or 0
            sum_plan_min += planned_min

            real_min = u.unit_duration_minutes() or 0
            sum_real_min += real_min

            qty_in = u.quantity_start or 0
            qty_out = u.quantity_end or 0
            loss = qty_in - qty_out
            loss_pct = round((loss / qty_in) * 100, 2) if qty_in > 0 else None

            sum_in += qty_in
            sum_out += qty_out

            ws.append([
                station.name,
                _to_local_str(u.start),
                _to_local_str(u.end),
                str(u.production_order.id_number) if u.production_order else "",
                u.sequence,
                planned_min,
                real_min,
                qty_in,
                qty_out,
                loss,
                loss_pct,
            ])

        # Podsumowanie
        total_loss = sum_in - sum_out
        total_loss_pct = round((total_loss / sum_in) * 100, 2) if sum_in > 0 else None

        ws.append([])
        ws.append([
            "RAZEM", "", "", "", "",
            sum_plan_min, sum_real_min, sum_in, sum_out, total_loss, total_loss_pct
        ])

        # Auto szerokość + autofilter
        ws.auto_filter.ref = f"A1:K{ws.max_row}"
        for col_idx, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        filename = f"straty_{station.name}_{year}-{month:02d}.xlsx".replace(" ", "_")
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
