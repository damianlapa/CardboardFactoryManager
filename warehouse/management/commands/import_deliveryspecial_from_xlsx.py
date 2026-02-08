from decimal import Decimal
from pathlib import Path
import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from warehouse.models import DeliverySpecial, DeliverySpecialItem


class Command(BaseCommand):
    help = "Importuje DeliverySpecial + itemy z pliku XLSX (bez zmian w logice przyjęć)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Ścieżka do pliku XLSX, np. /mnt/data/Zeszyt2.xlsx albo lokalnie: ./Zeszyt2.xlsx",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Nazwa arkusza (opcjonalnie). Jeśli puste, bierze pierwszy.",
        )
        parser.add_argument(
            "--ds-name",
            default="IMPORT GOTOWE 2026-01-01",
            help="Unikalna nazwa DeliverySpecial (domyślnie: IMPORT GOTOWE 2026-01-01)",
        )
        parser.add_argument(
            "--provider",
            default="IMPORT",
            help="Pole tekstowe provider w DeliverySpecial (domyślnie: IMPORT)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Nie zapisuje zmian, tylko pokazuje co by zrobił.",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        file_path = Path(opts["file"])
        if not file_path.exists():
            raise CommandError(f"Nie znaleziono pliku: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = opts["sheet"] or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Brak arkusza '{sheet_name}'. Dostępne: {wb.sheetnames}")

        ws = wb[sheet_name]

        ds_date = datetime.date(2026, 1, 1)
        ds_name = (opts["ds_name"] or "").strip()
        if not ds_name:
            raise CommandError("--ds-name nie może być pusty")

        provider = (opts["provider"] or "").strip() or "IMPORT"
        dry = opts["dry_run"]

        ds, created = DeliverySpecial.objects.get_or_create(
            name=ds_name,
            defaults={
                "provider": provider,
                "date": ds_date,
                "description": f"Import z XLSX: {file_path.name}",
                "processed": False,
            },
        )

        # jeżeli istnieje, upewnij się, że data jest taka jak chcemy
        if ds.date != ds_date:
            raise CommandError(
                f"DeliverySpecial '{ds.name}' już istnieje, ale ma datę {ds.date} (oczekiwano {ds_date})."
            )

        self.stdout.write(self.style.SUCCESS(
            f"DeliverySpecial: name='{ds.name}' date={ds.date} created={created}"
        ))

        created_count = 0
        updated_count = 0
        skipped_count = 0

        # Oczekiwany układ wiersza (bez nagłówków):
        # A: dimensions, B: customer/name, C: qty, D: notes, E: price
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue

            dims = (str(row[0]).strip() if row[0] is not None else "")
            flute = (str(row[1]).strip() if row[1] is not None else "")
            customer = (str(row[2]).strip() if row[2] is not None else "")
            qty_raw = row[3]
            notes = (str(row[4]).strip() if row[4] is not None else "")
            price_raw = row[5]

            # pomiń puste wiersze
            if not (dims or customer or qty_raw or notes or price_raw):
                skipped_count += 1
                continue

            # ilość
            try:
                qty = int(qty_raw or 0)
            except Exception:
                raise CommandError(f"Wiersz {idx}: niepoprawna ilość: {qty_raw}")
            if qty <= 0:
                skipped_count += 1
                continue

            # cena
            try:
                price = Decimal(str(price_raw or "0"))
            except Exception:
                raise CommandError(f"Wiersz {idx}: niepoprawna cena: {price_raw}")

            # nazwa itemu (max_length=64 w modelu)
            # format: "CUSTOMER | DIMS | NOTES"
            name = " | ".join([p for p in [customer.upper(), flute.upper(), dims, notes.upper()] if p]).strip()
            if len(name) > 64:
                name = name[:64].rstrip()

            if dry:
                self.stdout.write(f"[DRY] row={idx} -> item name='{name}' qty={qty} price={price}")
                continue

            item, it_created = DeliverySpecialItem.objects.get_or_create(
                delivery=ds,
                name=name,
                defaults={
                    "quantity": qty,
                    "price": price,
                    "processed": False,
                },
            )

            if it_created:
                created_count += 1
            else:
                # nadpisz quantity/price, żeby import był deterministyczny
                item.quantity = qty
                item.price = price
                item.processed = False
                item.save(update_fields=["quantity", "price", "processed"])
                updated_count += 1

        if dry:
            raise CommandError("DRY-RUN zakończony — transakcja przerwana celowo (nic nie zapisano).")

        self.stdout.write(self.style.SUCCESS(
            f"OK. Items: created={created_count}, updated={updated_count}, skipped={skipped_count}"
        ))
