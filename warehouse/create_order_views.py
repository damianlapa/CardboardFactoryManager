from django.views import View
from django.http import HttpResponse
from openpyxl import load_workbook
from io import BytesIO
import requests
import datetime
from oauth2client.service_account import ServiceAccountCredentials
import os
import gspread
from warehouse.models import Order


def get_gluer_number(dimensions, customer=None):
    params = {
        'dimensions': dimensions,
        'customer': customer
    }

    response = requests.get('https://paker-wroclaw.herokuapp.com/gluernumberget/', params=params)
    # response = requests.get('http://127.0.0.1:8000/gluernumberget/', params=params)

    # Sprawdzenie statusu odpowiedzi
    if response.status_code == 200:
        data = response.json()  # zakładając, że odpowiedź jest w formacie JSON
        return data


def get_polymer_number(name=None, dimensions=None, customer=None):
    params = {
        'name': name,
    }

    response = requests.get('https://paker-wroclaw.herokuapp.com/polymernumberget/', params=params)

    # Sprawdzenie statusu odpowiedzi
    if response.status_code == 200:
        data = response.json()
        return data
    else:
        print(response.status_code)


def get_punch_number(dimensions=None, name=None):
    params = {
        'dimensions': dimensions,
        'name': name,
    }

    response = requests.get('https://paker-wroclaw.herokuapp.com/punchnumberget/', params=params)

    # Sprawdzenie statusu odpowiedzi
    if response.status_code == 200:
        data = response.json()
        return data['punch']
    else:
        print(response.status_code)


def get_data_by_values(provider: str, number: str, year_short: str, year_full: str = '2024'):
    GOOGLE_SHEETS_CREDENTIALS = {
        "type": os.environ['PE_TYPE'],
        "project_id": os.environ['PE_PROJECT_ID'],
        "private_key_id": os.environ['PE_PRIVATE_KEY_ID'],
        "private_key": os.environ['PE_PRIVATE_KEY'].replace('\\n', '\n'),
        "client_email": os.environ['PE_CLIENT_EMAIL'],
        "client_id": os.environ['PE_CLIENT_ID'],
        "auth_uri": os.environ['PE_AUTH_URI'],
        "token_uri": os.environ['PE_TOKEN_URI'],
        "auth_provider_x509_cert_url": os.environ['PE_AUTH_PROVIDER_X509_CERT_URL'],
        "client_x509_cert_url": os.environ['PE_CLIENT_X509_CERT_URL'],
    }

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(GOOGLE_SHEETS_CREDENTIALS, scope)
    client = gspread.authorize(creds)

    sheet = client.open("PAKER TEKTURA ZAMÓWIENIA").worksheet(f"ZAMÓWIENIA {year_full}")
    all_values = sheet.get_all_values()

    for row in all_values:
        if len(row) >= 3 and row[0] == provider and row[1] == number and row[2] == year_short:
            return row

    return None


class GenerateOrderInlineView(View):
    def get(self, request, order_id, *args, **kwargs):
        # Pobierz plik wzoru z linku
        url = "https://paker.eu/wp-content/uploads/2025/07/wzor2euro.xlsx"
        response = requests.get(url)
        order = Order.objects.get(id=order_id)
        num, year = order.order_id.split('/')
        order_provider = order.provider.shortcut

        data = get_data_by_values(order_provider, num, year, f'20{year}')
        
        if response.status_code != 200:
            return HttpResponse("Nie udało się pobrać wzoru", status=500)

        # Załaduj do workbooka i zmodyfikuj
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active

        dane_polimeru = get_polymer_number(data[10].strip())
        if dane_polimeru:
            ws['E14'] = dane_polimeru['number']
            ws['M14'] = dane_polimeru['colors']

        # INFORMACJE DODATKOWE
        ws['A29'] = data[25]
        numer_zlecenia = f'{data[0]} {data[1]}/{data[2]} {data[18]}'
        nazwa_zlecenia = f'{data[0]} {data[1]}_{data[2]}_{data[18]}'
        output_file_path = f'{DESTINATION}/zlecenie {nazwa_zlecenia}.xlsx'
        ws['J3'] = numer_zlecenia

        # DATY
        date = datetime.datetime.today()
        ws['W1'] = date
        if data[5]:
            try:
                date2 = datetime.datetime.strptime(data[5], '%Y-%m-%d').date() + datetime.timedelta(days=21)
                day = date2.day if len(str(date2.day)) == 2 else f'0{date2.day}'
                month = date2.month if len(str(date2.month)) == 2 else f'0{date2.month}'
                date2 = f'{day}-{month}-{date2.year}'
            except Exception as e:
                date2 = ''
            ws['W2'] = date2 if not data[7] else f'{data[7][8:]}-{data[7][5:7]}-{data[7][:4]}'

        # proces produkcyjny

        if data[9]:
            stanowiska = data[9].split('/')
            for num in range(len(stanowiska)):
                komorka = 5 + num * 3
                if stanowiska[num] == 'SKL':
                    numer_sklejarka = get_gluer_number(data[23].lower().strip(), data[18].strip().upper())
                    if numer_sklejarka:
                        ws[f'Q{komorka}'] = f'SKLEJARKA({numer_sklejarka["number"]})'
                        ws['A50'] = numer_sklejarka['comments']
                    else:
                        ws[f'Q{komorka}'] = f'SKLEJARKA(   )'
                else:
                    ws[f'Q{komorka}'] = stanowiska[num]

        # wymiary
        wymiary = data[23].lower().split('x')
        if len(wymiary) == 3:
            ws['A10'] = wymiary[0]
            ws['G10'] = wymiary[1]
            ws['M10'] = wymiary[2]
        elif len(wymiary) == 2:
            ws['A10'] = wymiary[0]
            ws['G10'] = wymiary[1]
        elif len(wymiary) == 1:
            ws['A10'] = wymiary[0]

        # oznaczenie
        oznaczenie = f'{data[19]}{data[20]}{data[21]}'
        ws['A23'] = oznaczenie

        # format tektury
        format_tektury = f'{data[12]}x{data[13]}'
        ws['F23'] = format_tektury

        # ilosc zamowiona
        ilosc_zamowiona = data[14]
        ws['L23'] = ilosc_zamowiona

        # ilosc przyjeta
        ilosc_przyjeta = data[15]
        ws['N23'] = ilosc_przyjeta

        # OZNACZENIE ETYKIETY
        oznaczenie_etykiety = data[24]
        ws['K62'] = oznaczenie_etykiety

        if 'TYG' in data[9]:
            if data[11]:
                wykrojnik = get_punch_number(name=data[11].strip())
            else:
                wykrojnik = None
            if wykrojnik:
                ws['E17'] = wykrojnik
            else:
                print('BRAK WYKROJNIKA! -> SPRAWDŹ POPRAWNOŚĆ')

        # Zapisz do pamięci (RAM)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Zwróć jako odpowiedź
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'inline; filename="zlecenie_test.xlsx"'
        return response
